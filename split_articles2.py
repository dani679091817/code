import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

INPUT_FILE = "LISTE ARTICLE.xlsx"
OUTPUT_PREFIX = "liste_article2"
MAX_ENTRIES = 2000

wb_in = openpyxl.load_workbook(INPUT_FILE)
ws_in = wb_in.active

# Row 7 is the header row; data starts from row 8
# Column A (index 0): Référence
# Column D (index 3): Désignation
# Column R (index 17): Prix
data_rows = []
for row in ws_in.iter_rows(min_row=8, values_only=True):
    if any(v is not None for v in row):
        ref = row[0]          # column A: Référence
        designation = row[3]  # column D: Désignation
        prix = row[17]        # column R: Prix
        # Format price: convert whole-number floats to int (no trailing .00)
        if prix is not None:
            if isinstance(prix, float) and prix == int(prix):
                prix = int(prix)
        data_rows.append((ref, designation, prix))

headers = ("Référence", "Désignation", "Prix")

num_files = (len(data_rows) + MAX_ENTRIES - 1) // MAX_ENTRIES

for file_index in range(num_files):
    start = file_index * MAX_ENTRIES
    end = start + MAX_ENTRIES
    chunk = data_rows[start:end]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Liste d'articles"

    # Write headers
    ws_out.append(headers)
    for cell in ws_out[1]:
        cell.font = Font(bold=True)

    # Write data
    for row in chunk:
        ws_out.append(row)

    # Auto-fit column widths (approximate)
    for col_idx, col_cells in enumerate(ws_out.columns, 1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
        ws_out.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    output_filename = f"{OUTPUT_PREFIX}_{file_index + 1}.xlsx"
    wb_out.save(output_filename)
    print(f"Saved {output_filename} with {len(chunk)} entries")

print("Done.")
