import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "liste_article.xlsx"
OUTPUT_PREFIX = "liste_article"
MAX_ENTRIES = 2000
NUM_FILES = 5

wb_in = openpyxl.load_workbook(INPUT_FILE)
ws_in = wb_in.active

# Collect data rows (starting from row 3, skipping title row 1 and header row 2)
data_rows = []
for row in ws_in.iter_rows(min_row=3, values_only=True):
    if any(v is not None for v in row):
        ref = row[0]
        designation = row[1]
        prix = row[3]
        # Format price: remove trailing .0 for whole numbers
        if prix is not None:
            if isinstance(prix, float) and prix == int(prix):
                prix = int(prix)
        data_rows.append((ref, designation, prix))

headers = ("Référence", "Désignation", "Prix TTC")

for file_index in range(NUM_FILES):
    start = file_index * MAX_ENTRIES
    end = start + MAX_ENTRIES
    chunk = data_rows[start:end]
    if not chunk:
        break

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Liste d'articles"

    # Write headers
    ws_out.append(headers)
    header_row = ws_out[1]
    for cell in header_row:
        cell.font = Font(bold=True)

    # Write data
    for row in chunk:
        ws_out.append(row)

    # Auto-fit column widths (approximate)
    for col_idx, col_cells in enumerate(ws_out.columns, 1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
        ws_out.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    output_filename = f"{OUTPUT_PREFIX}_{file_index + 1}.xlsx"
    wb_out.save(output_filename)
    print(f"Saved {output_filename} with {len(chunk)} entries")

print("Done.")
